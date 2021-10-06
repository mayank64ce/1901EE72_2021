import os
import csv
import openpyxl

SUBJECT_FILE = 'subjects_master.csv'
STUDENT_FILE = 'names-roll.csv'
GRADES = 'grades.csv'
OUTPUT_DIR = 'output'
CWD = os.path.dirname(os.path.realpath(__file__))
subjects = {}
students = {}


def get_subjects():

    global CWD
    file_path = os.path.join(CWD, SUBJECT_FILE)
    with open(file_path, newline='') as subject_data:
        rows = csv.reader(subject_data, delimiter=',', quotechar='"')
        next(rows)
        for row in rows:
            subjects[row[0]] = {
                "subname": row[1],
                "ltp": row[2],
                "crd": row[3]
            }


def get_students():
    global CWD
    file_path = os.path.join(CWD, STUDENT_FILE)
    with open(file_path, newline='') as student_data:
        rows = csv.reader(student_data, delimiter=',', quotechar='|')
        next(rows)
        for row in rows:
            students[row[0]] = {
                "name": row[1]
            }
            # break


def setup_directories():
    global CWD
    if not os.path.isdir(os.path.join(CWD, OUTPUT_DIR)):
        os.mkdir(os.path.join(CWD, OUTPUT_DIR))


def generate_semwise():
    global CWD, OUTPUT_DIR
    output_dir = os.path.join(CWD, OUTPUT_DIR)
    with open(os.path.join(CWD, GRADES)) as grade_master:
        rows = csv.reader(grade_master, delimiter=',', quotechar='|')
        next(rows)
        for row in rows:
            rollno = row[0]
            sem = row[1]
            subcode = row[2]
            credit = row[3]
            grade = row[4].strip()
            sub_type = row[5]
            output_file = rollno + ".xlsx"
            output_file = os.path.join(output_dir, output_file)
            sheet_name = "Sem " + str(sem).strip()

            if not os.path.exists(output_file):
                wb = openpyxl.Workbook()
            else:
                wb = openpyxl.load_workbook(output_file)

            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
                ws = wb[sheet_name]
                # print("bug")
                ws.append(["SI No.", "Subject No.", "Subject Name",
                           "L-T-P", "Credit", "Subject Type", "Grade"])

            ws = wb[sheet_name]

            data = [ws.max_row]
            data.append(subcode)
            data.append(subjects[subcode]["subname"])
            data.append(subjects[subcode]["ltp"])
            data.append(credit)
            data.append(sub_type)
            data.append(grade)

            ws.append(data)
            wb.save(filename=os.path.join(output_dir, output_file))
    return


def generate_overall():

    grade_map = {}

    grade_map["AA"] = 10
    grade_map["AB"] = 9
    grade_map["BB"] = 8
    grade_map["BC"] = 7
    grade_map["CC"] = 6
    grade_map["CD"] = 5
    grade_map["DD"] = 4
    grade_map["F"] = 0
    grade_map["I"] = 0

    grade_map["AA*"] = 10
    grade_map["AB*"] = 9
    grade_map["BB*"] = 8
    grade_map["BC*"] = 7
    grade_map["CC*"] = 6
    grade_map["CD*"] = 5
    grade_map["DD*"] = 4
    grade_map["F*"] = 0
    grade_map["I*"] = 0

    global CWD, OUTPUT_DIR
    output_dir = os.path.join(CWD, OUTPUT_DIR)
    for ROLLNO in students.keys():
        file = ROLLNO + ".xlsx"

        rollno = ['Roll No.']
        name = ['Name']
        discipline = ['Discipline']  # need to figure out discipline
        semesters = ['Semester No.']
        semwise_credit = ['Semester wise Credit Taken']
        spi = ['SPI']
        total_credits = ['Total Credits Taken']
        cpi = ['CPI']

        rollno.append(ROLLNO)
        name.append(students[ROLLNO]['name'])
        discipline.append(ROLLNO[4:6])

        output_file = os.path.join(output_dir, file)
        if not os.path.exists(output_file):
            continue
        else:
            wb = openpyxl.load_workbook(output_file)

        if 'Sheet' in wb.sheetnames:
            wb['Sheet'].title = 'Overall'
        total_sems = len(wb.sheetnames)

        credits_taken = 0
        total_cpi = 0
        for sem in wb.sheetnames:
            if sem == 'Overall':
                continue
            sheet = wb[sem]
            current_credits = 0
            current_weighted_sum = 0

            rows = sheet.max_row

            for i in range(2, rows+1):
                current_credits += int(sheet.cell(i, 5).value)
                current_weighted_sum += grade_map[sheet.cell(
                    i, 7).value] * int(sheet.cell(i, 5).value)

            current_spi = current_weighted_sum / current_credits
            current_spi = round(current_spi, 2)
            total_cpi = (total_cpi * credits_taken + current_spi *
                         current_credits) / (credits_taken + current_credits)
            total_cpi = round(total_cpi, 2)
            credits_taken += current_credits

            semesters.append(sem[4:])
            semwise_credit.append(current_credits)
            spi.append(current_spi)
            total_credits.append(credits_taken)
            cpi.append(total_cpi)

        wb['Overall'].append(rollno)
        wb['Overall'].append(name)
        wb['Overall'].append(discipline)
        wb['Overall'].append(semesters)
        wb['Overall'].append(semwise_credit)
        wb['Overall'].append(spi)
        wb['Overall'].append(total_credits)
        wb['Overall'].append(cpi)

        wb.save(filename=output_file)
    return


setup_directories()
get_subjects()
get_students()
generate_semwise()
generate_overall()
