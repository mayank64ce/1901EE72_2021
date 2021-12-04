import pandas as pd
import json
import os
import shutil
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
import openpyxl
from .validate import check
from subprocess import run

BASE_PATH = os.path.dirname(
    os.path.realpath(__file__))

OUTPUT_FILE_PATH = BASE_PATH + "/output/rollNumberWise"
RESPONSE_FILE = BASE_PATH + "/uploads/csv/responses.csv"
MASTER_FILE = BASE_PATH + "/uploads/csv/master.csv"
TEMPLATE_FILE = BASE_PATH + "/templates/marksheet_template.xlsx"

# num_questions = 0

stud_info = {}
answers = {}
ans_idx = check()
coords = {
    "name": "B6",
    "rollnumber": "B7",
    "#_right": "B10",
    "#_wrong": "C10",
    "#_left": "D10",
    "pos": "B11",
    "neg": "C11",
    "max": "E10",
    "total": "B12",
    "wrong": "C12",
    "final": "E12"
}

# print(check())


def process_master():
    # pass
    df = pd.read_csv(MASTER_FILE)
    for i in range(len(df)):
        rollno = df.loc[i, "roll"]
        name = df.loc[i, "name"]
        stud_info[rollno] = {
            "name": name,
            "marks": "",
            "status": []
        }


def get_answer():
    df = pd.read_csv(RESPONSE_FILE)
    # print(df.columns)
    num_questions = len(df.columns) - 7
    return num_questions


def generate_marksheet():
    num_questions = get_answer()
    start = 7
    df = pd.read_csv(RESPONSE_FILE)

    red_font = Font(color="E41B17")
    blue_font = Font(color="1589FF")
    green_font = Font(color="228B22")

    for i in range(len(df)):
        rollno = df.loc[i, "Roll Number"]
        # if rollno == "ANSWER":
        #     continue
        filename = rollno.upper() + ".xlsx"
        target_path = OUTPUT_FILE_PATH + "/" + filename

        shutil.copyfile(TEMPLATE_FILE, target_path)

        wb = openpyxl.load_workbook(target_path)
        ws = wb['Sheet1']
        from .views import negative, positive

        correct = 0
        wrong = 0
        left = 0
        begin = 16
        for j in range(num_questions):
            studCellNo = "A"+str(begin+j)
            ansCellNo = "B"+str(begin+j)
            ws[studCellNo].value = df.loc[i, "Unnamed: "+str(start+j)]
            ws[ansCellNo].value = df.loc[ans_idx, "Unnamed: " + str(start+j)]
            ws[ansCellNo].font = blue_font

            if ws[studCellNo].value == ws[ansCellNo].value:
                correct += 1
                ws[studCellNo].font = green_font
                # print("*******************",
                #       ws[studCellNo].value, ws[ansCellNo].value, "green")
            elif not ws[studCellNo].value:
                left += 1
                # print("*******************",
                #       ws[studCellNo].value, ws[ansCellNo].value, "no color")
            elif ws[studCellNo].value != ws[ansCellNo].value:
                wrong += 1
                ws[studCellNo].font = red_font
                # print("*******************",
                #   ws[studCellNo].value, ws[ansCellNo].value, "red")

        ws.title = 'quiz'
        ws[coords["name"]].value = stud_info[rollno]["name"]
        ws[coords["rollnumber"]].value = rollno

        ws[coords["#_right"]].value = correct
        ws[coords["#_left"]].value = left
        ws[coords["#_wrong"]].value = wrong

        ws[coords["pos"]].value = positive
        ws[coords["neg"]].value = negative
        ws[coords["max"]].value = num_questions
        ws[coords["total"]].value = positive * correct
        ws[coords["wrong"]].value = negative * wrong

        ws[coords["final"]].value = str(
            positive * correct + negative * wrong) + "/" + str(positive * num_questions)

        stud_info[rollno]["status"] = [correct, wrong, left]
        stud_info[rollno]["marks"] = str(
            positive * correct + negative * wrong) + "/" + str(positive * num_questions)

        wb.save(target_path)
        # print(negative, positive)
        # return


def generate_marksheet_for_absentees():

    for rollno in stud_info:
        if len(stud_info[rollno]["marks"]) != 0:
            continue
        # create marksheet
        filename = rollno.upper() + ".xlsx"
        target_path = OUTPUT_FILE_PATH + "/" + filename

        shutil.copyfile(TEMPLATE_FILE, target_path)

        wb = openpyxl.load_workbook(target_path)
        ws = wb['Sheet1']
        from .views import negative, positive

        ws.title = 'quiz'
        ws[coords["name"]].value = stud_info[rollno]["name"]
        ws[coords["rollnumber"]].value = rollno

        ws[coords["pos"]].value = positive
        ws[coords["neg"]].value = negative

        ws[coords["#_right"]].value = 0
        ws[coords["#_left"]].value = 0
        ws[coords["#_wrong"]].value = 0

        ws[coords["final"]].value = "Absent"
        wb.save(target_path)

    return


# get_answer()
def driver():
    print("Generating All Marksheets.....")
    process_master()
    get_answer()
    generate_marksheet()
    generate_marksheet_for_absentees()
    
    run(f'rm {OUTPUT_FILE_PATH}/placeholder.txt', shell=True)

    with open(BASE_PATH + "/stud_info.json", "w") as outfile:
        json.dump(stud_info, outfile)
    print("All Marksheets Generated !!")


# driver()
