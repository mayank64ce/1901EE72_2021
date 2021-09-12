"""
@author: mayank64ce
This program categorizes Course information according to Students and Subjects
"""
import os
import csv
import openpyxl

csv_file = 'regtable_old.csv'
DIR_ROLL_NUMBER = 'output_individual_roll'
DIR_SUBJECT = 'output_by_subject'


def output_by_subject():
    """
        This methods iterates over the given course dataset
        and puts together all students who took a particular course.
    """
    # print('output_by_subject')
    # fo = open(csv_file)
    with open(csv_file, newline='') as csvfile:
        lines = csv.reader(csvfile, delimiter=',', quotechar='|')
        next(lines)
        for line in lines:
            subno = line[3]
            file_path = os.path.join(DIR_SUBJECT, subno + '.xlsx')
            # print(file_path)

            if not os.path.exists(file_path):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["rollno", "register_sem", "subno", "sub_type"])
            else:
                wb = openpyxl.load_workbook(file_path)

            ws = wb.active

            output = [line[0], line[1], line[3], line[8]]
            ws.append(output)
            wb.save(filename=file_path)
            # if i == 5:
            #     break
    return


def output_individual_roll():
    # print('output_individual_roll')
    """
        This methods iterates over the given course dataset
        and puts together all courses taken a particular students.
    """
    with open(csv_file, newline='') as csvfile:
        lines = csv.reader(csvfile, delimiter=',', quotechar='|')
        next(lines)
        for line in lines:
            rollno = line[0]
            file_path = os.path.join(DIR_ROLL_NUMBER, rollno + '.xlsx')
            # print(file_path)

            if not os.path.exists(file_path):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["rollno", "register_sem", "subno", "sub_type"])
            else:
                wb = openpyxl.load_workbook(file_path)

            ws = wb.active

            output = [line[0], line[1], line[3], line[8]]
            ws.append(output)
            wb.save(filename=file_path)
            # if i == 5:
            #     break
    return


def setup_directories():
    """
        Sets up the output directories
    """
    cwd = os.getcwd()
    if not os.path.isdir(os.path.join(cwd, DIR_ROLL_NUMBER)):
        os.mkdir(os.path.join(cwd, DIR_ROLL_NUMBER))
    if not os.path.isdir(os.path.join(cwd, DIR_SUBJECT)):
        os.mkdir(os.path.join(cwd, DIR_SUBJECT))


# Driver Code
setup_directories()
output_individual_roll()
output_by_subject()
